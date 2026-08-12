"""Warstwa danych: SQLAlchemy (SQLite lokalnie / Postgres na produkcji).

Ustaw DATABASE_URL (np. z Supabase/Neon), aby użyć Postgresa i mieć trwałe dane.
Bez DATABASE_URL używany jest lokalny plik SQLite (data/ondre.db) — do testów.
"""
import json
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from sqlalchemy import (BigInteger, Column, Float, Integer, LargeBinary,
                        MetaData, String, Table, Text, create_engine, delete,
                        func, insert, select, text, update)

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
CARDS_DIR = DATA_DIR / "cards"        # cache tymczasowy (bajty kart są w bazie)
OFFERS_DIR = DATA_DIR / "offers"      # cache tymczasowy ofert do pobrania
SEED_XLSX = ROOT / "assets" / "seed" / "Cennik_znormalizowany_v1.xlsx"

TIERS = [
    ("Katalogowa", "m_katalog"),
    ("Klient stały", "m_staly"),
    ("Pośrednik", "m_posrednik"),
    ("Agencyjny", "m_agencyjny"),
    ("Jedi", "m_jedi"),
]
TIER_NAMES = [t[0] for t in TIERS]
TIER_COL = dict(TIERS)

DEFAULT_SETTINGS = {
    "company_name": "ONDRE Sp. z o.o. Sp. k.",
    "company_address": "ul. Składowa 3, 62-081 Przeźmierowo",
    "company_nip": "7811952034",
    "company_regon": "368103366",
    "company_email": "biuro@ondre.pl",
    "company_phone": "+48 61 666 04 01",
    "company_web": "www.ondre.pl",
    "company_bank": "23 8362 1044 5508 0259 2000 0010",
    "offer_prefix": "OF",
    "offer_validity_days": "14",
    "vat_rate": "23",
    "payment_terms": "Płatność: przelew. Podane ceny są cenami netto — zostanie doliczony podatek VAT.",
    "pipedrive_tier_field": "",
}


# --------------------------------------------------------------------------- #
#  Silnik / definicje tabel
# --------------------------------------------------------------------------- #
def _secret(name, default=""):
    """DATABASE_URL i inne sekrety: najpierw env, potem st.secrets."""
    if name in os.environ:
        return os.environ[name]
    try:
        import streamlit as st
        if name in st.secrets:
            return st.secrets[name]
    except Exception:
        pass
    return default


def _db_url() -> str:
    url = _secret("DATABASE_URL", "").strip()
    if not url:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        return "sqlite:///%s" % (DATA_DIR / "ondre.db")
    # normalizacja do psycopg3
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql+psycopg://", 1)
    elif url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+psycopg://", 1)
    return url


_ENGINE = None
_INIT_DONE = False
md = MetaData()

products = Table(
    "products", md,
    Column("id", String(16), primary_key=True),
    Column("section", String(128)), Column("name", Text), Column("variant", String(64)),
    Column("unit", String(8), default="m2"), Column("base_cost", Float),
    Column("m_katalog", Float), Column("m_staly", Float), Column("m_posrednik", Float),
    Column("m_agencyjny", Float), Column("m_jedi", Float),
    Column("active", Integer, default=1), Column("card_file", String(128)),
    Column("note", Text), Column("min_price", Float),
    Column("scope", String(16), default="ondre"), Column("client_key", String(200)),
    Column("status", String(16), default="active"), Column("created_by", String(128)),
    Column("def_szer", Float), Column("def_wys", Float),
)
price_history = Table(
    "price_history", md,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("ts", String(32)), Column("usr", String(128)), Column("product_id", String(16)),
    Column("field", String(64)), Column("old_value", Text), Column("new_value", Text),
)
settings_t = Table(
    "settings", md,
    Column("key", String(64), primary_key=True), Column("value", Text),
)
offers = Table(
    "offers", md,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("number", String(48)), Column("ts", String(32)),
    Column("user_id", Integer), Column("user_name", String(128)),
    Column("client_name", Text), Column("client_json", Text), Column("tier", String(32)),
    Column("items_json", Text), Column("total_net", Float),
    Column("pdf_name", String(160)), Column("pdf_blob", LargeBinary),
    Column("pipedrive_deal_id", BigInteger),
)
cards = Table(
    "cards", md,
    Column("product_id", String(16), primary_key=True),
    Column("filename", String(160)), Column("data", LargeBinary),
    Column("uploaded_by", String(128)), Column("ts", String(32)),
)
users = Table(
    "users", md,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("email", String(160), unique=True), Column("name", String(128)),
    Column("phone", String(48)), Column("password_hash", Text),
    Column("role", String(16), default="handlowiec"), Column("active", Integer, default=1),
    Column("created_ts", String(32)), Column("pipedrive_token", Text),
)


def get_engine():
    global _ENGINE
    if _ENGINE is None:
        url = _db_url()
        kw = {"pool_pre_ping": True, "future": True}
        if url.startswith("sqlite"):
            kw["connect_args"] = {"check_same_thread": False}
        _ENGINE = create_engine(url, **kw)
    return _ENGINE


def init_db():
    global _INIT_DONE
    if _INIT_DONE:
        return
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    CARDS_DIR.mkdir(parents=True, exist_ok=True)
    OFFERS_DIR.mkdir(parents=True, exist_ok=True)
    md.create_all(get_engine())
    _migrate()
    _INIT_DONE = True


def _migrate():
    """Dodaje brakujące kolumny w istniejących tabelach + wypełnia domyślne."""
    from sqlalchemy import inspect, text
    eng = get_engine()
    try:
        cols = [c["name"] for c in inspect(eng).get_columns("products")]
    except Exception:
        return
    new_cols = {
        "min_price": "FLOAT", "scope": "VARCHAR(16)", "client_key": "VARCHAR(200)",
        "status": "VARCHAR(16)", "created_by": "VARCHAR(128)",
        "def_szer": "FLOAT", "def_wys": "FLOAT",
    }
    for name, typ in new_cols.items():
        if name not in cols:
            try:
                with eng.begin() as c:
                    c.execute(text("ALTER TABLE products ADD COLUMN %s %s" % (name, typ)))
            except Exception:
                pass
    # backfill wartości domyślnych (jednorazowo — tylko puste)
    try:
        with eng.begin() as c:
            c.execute(text("UPDATE products SET scope='ondre' WHERE scope IS NULL"))
            c.execute(text("UPDATE products SET status='active' WHERE status IS NULL"))
            c.execute(text("UPDATE products SET min_price=150 WHERE min_price IS NULL"))
    except Exception:
        pass
    # stałe wymiary konstrukcji — sparsuj z nazwy „WxH" (cm) tam, gdzie brak (jednorazowo)
    try:
        import re
        flag = get_settings().get("migr_defdims")
        if not flag:
            constr = ("Roll-upy", "LED BOX / standy LED", "Ścianki LED", "X-bannery")
            with eng.begin() as c:
                rows = c.execute(text("SELECT id,name,section FROM products "
                                      "WHERE def_szer IS NULL")).fetchall()
                for pid, name, section in rows:
                    if section in constr and name:
                        m = re.search(r"(\d{2,4})\s*[x×]\s*(\d{2,4})", name)
                        if m:
                            w, h = int(m.group(1)) / 100.0, int(m.group(2)) / 100.0
                            c.execute(text("UPDATE products SET def_szer=:w, def_wys=:h "
                                           "WHERE id=:i"), {"w": w, "h": h, "i": pid})
            save_settings({"migr_defdims": "1"})
    except Exception:
        pass
    # kolumny w tabeli users
    try:
        ucols = [c["name"] for c in inspect(eng).get_columns("users")]
        if "pipedrive_token" not in ucols:
            with eng.begin() as c:
                c.execute(text("ALTER TABLE users ADD COLUMN pipedrive_token TEXT"))
    except Exception:
        pass


def is_postgres() -> bool:
    return get_engine().dialect.name == "postgresql"


# --------------------------------------------------------------------------- #
#  Produkty / cennik
# --------------------------------------------------------------------------- #
def seed_if_empty() -> bool:
    init_db()
    with get_engine().begin() as c:
        n = c.execute(select(func.count()).select_from(products)).scalar()
    if n == 0 and SEED_XLSX.exists():
        import_xlsx(SEED_XLSX, user="seed")
        return True
    return False


def _upsert_product(conn, vals: dict):
    """Wstaw-lub-zaktualizuj po id (odporne na powtórny import / kolizje ID)."""
    setvals = {k: v for k, v in vals.items() if k != "id"}
    name = conn.dialect.name
    if name == "postgresql":
        from sqlalchemy.dialects.postgresql import insert as _ins
    elif name == "sqlite":
        from sqlalchemy.dialects.sqlite import insert as _ins
    else:
        exists = conn.execute(select(products.c.id).where(products.c.id == vals["id"])).first()
        if exists:
            conn.execute(update(products).where(products.c.id == vals["id"]).values(setvals))
        else:
            conn.execute(insert(products).values(**vals))
        return
    stmt = _ins(products).values(**vals).on_conflict_do_update(
        index_elements=["id"], set_=setvals)
    conn.execute(stmt)


def import_xlsx(path, user="import") -> int:
    from openpyxl import load_workbook
    init_db()
    wb = load_workbook(path, data_only=True)
    ws = wb["Cennik_produkty"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    count = 0
    with get_engine().begin() as c:
        for r in rows:
            if not r or not r[0]:
                continue
            cost = r[5]
            cost = float(cost) if cost not in (None, "", 0) else None
            marks = [float(m) if m not in (None, "") else None for m in r[6:11]]
            card = r[16] if len(r) > 16 else None
            _upsert_product(c, dict(
                id=r[0], section=r[1], name=r[2], variant=r[3], unit=r[4] or "m2",
                base_cost=cost, m_katalog=marks[0], m_staly=marks[1], m_posrednik=marks[2],
                m_agencyjny=marks[3], m_jedi=marks[4], active=1, card_file=card,
                scope="ondre", status="active", min_price=150))
            count += 1
        c.execute(insert(price_history).values(
            ts=_now(), usr=user, product_id="*", field="import_xlsx",
            old_value="", new_value=str(count)))
    return count


def products_df(active_only=False, scope=None, status=None, client_key=None) -> pd.DataFrame:
    init_db()
    q = ("SELECT id,section,name,variant,unit,base_cost,m_katalog,m_staly,"
         "m_posrednik,m_agencyjny,m_jedi,active,card_file,note,min_price,"
         "scope,client_key,status,created_by,def_szer,def_wys FROM products")
    conds, params = [], {}
    if active_only:
        conds.append("active=1")
    if scope is not None:
        conds.append("scope = :scope"); params["scope"] = scope
    if status is not None:
        conds.append("status = :status"); params["status"] = status
    if client_key is not None:
        conds.append("client_key = :ck"); params["ck"] = client_key
    if conds:
        q += " WHERE " + " AND ".join(conds)
    q += " ORDER BY id"
    with get_engine().connect() as c:
        return pd.read_sql_query(text(q), c, params=params)


def catalog_products() -> pd.DataFrame:
    """Aktywne produkty ONDRE (do cennika i wyboru w ofercie)."""
    return products_df(active_only=True, scope="ondre", status="active")


def client_products(client_key) -> pd.DataFrame:
    """Aktywne produkty przypisane do danego klienta."""
    if not client_key:
        return products_df().iloc[0:0]
    return products_df(status="active", scope="client", client_key=client_key)


def pending_products() -> pd.DataFrame:
    return products_df(status="pending")


def count_pending() -> int:
    init_db()
    with get_engine().connect() as c:
        return c.execute(select(func.count()).select_from(products)
                         .where(products.c.status == "pending")).scalar() or 0


def create_custom_product(name, section, unit, base_cost, marks, min_price,
                          scope="ondre", client_key=None, status="active",
                          created_by="", variant="", def_szer=None, def_wys=None) -> str:
    """Tworzy produkt (ONDRE / klienta / oczekujący). marks: 5 narzutów."""
    pid = next_product_id()
    with get_engine().begin() as c:
        c.execute(insert(products).values(
            id=pid, section=section, name=name, variant=variant, unit=unit,
            base_cost=base_cost, m_katalog=marks[0], m_staly=marks[1], m_posrednik=marks[2],
            m_agencyjny=marks[3], m_jedi=marks[4], active=1, min_price=min_price,
            scope=scope, client_key=client_key, status=status, created_by=created_by,
            def_szer=def_szer, def_wys=def_wys))
        c.execute(insert(price_history).values(
            ts=_now(), usr=created_by or "system", product_id=pid,
            field="create_%s" % status, old_value="", new_value=name))
    return pid


def approve_product(pid, user="admin"):
    with get_engine().begin() as c:
        c.execute(update(products).where(products.c.id == pid).values(status="active"))
        c.execute(insert(price_history).values(
            ts=_now(), usr=user, product_id=pid, field="approve", old_value="pending",
            new_value="active"))


def reject_product(pid, user="admin"):
    with get_engine().begin() as c:
        c.execute(delete(products).where(products.c.id == pid))
        c.execute(insert(price_history).values(
            ts=_now(), usr=user, product_id=pid, field="reject", old_value="pending",
            new_value="deleted"))


def sections_list() -> list:
    df = products_df()
    return sorted(x for x in df["section"].dropna().unique())


def save_products_df(df: pd.DataFrame, user="admin") -> int:
    old = products_df().set_index("id")
    ts = _now()
    cols = ["section", "name", "variant", "unit", "base_cost",
            "m_katalog", "m_staly", "m_posrednik", "m_agencyjny", "m_jedi",
            "active", "note", "min_price"]
    changes = 0
    with get_engine().begin() as c:
        for _, row in df.iterrows():
            pid = row["id"]
            if pid in old.index:
                for col in cols:
                    ov, nv = old.loc[pid, col], row.get(col)
                    ov = None if pd.isna(ov) else ov
                    nv = None if pd.isna(nv) else nv
                    if ov != nv:
                        c.execute(update(products).where(products.c.id == pid).values({col: nv}))
                        c.execute(insert(price_history).values(
                            ts=ts, usr=user, product_id=pid, field=col,
                            old_value=str(ov), new_value=str(nv)))
                        changes += 1
            else:
                vals = {col: (None if pd.isna(row.get(col)) else row.get(col)) for col in cols}
                vals["id"] = pid
                _upsert_product(c, vals)
                c.execute(insert(price_history).values(
                    ts=ts, usr=user, product_id=pid, field="create",
                    old_value="", new_value=str(row.get("name", ""))))
                changes += 1
    return changes


def next_product_id() -> str:
    df = products_df()
    nums = [int(x[1:]) for x in df["id"] if str(x).startswith("P") and str(x)[1:].isdigit()]
    return "P%03d" % (max(nums) + 1 if nums else 1)


def history_df(limit=100) -> pd.DataFrame:
    with get_engine().connect() as c:
        return pd.read_sql_query(
            text("SELECT ts,usr,product_id,field,old_value,new_value FROM price_history "
                 "ORDER BY id DESC LIMIT :l"), c, params={"l": limit})


def export_xlsx(path):
    df = products_df()
    df.to_excel(path, sheet_name="Cennik_produkty_export", index=False)
    return path


# --------------------------------------------------------------------------- #
#  Karty produktowe (bajty w bazie)
# --------------------------------------------------------------------------- #
def set_card(product_id, filename, data: bytes, user="admin"):
    init_db()
    with get_engine().begin() as c:
        c.execute(delete(cards).where(cards.c.product_id == product_id))
        c.execute(insert(cards).values(product_id=product_id, filename=filename,
                                       data=data, uploaded_by=user, ts=_now()))
        c.execute(update(products).where(products.c.id == product_id)
                  .values(card_file=filename))


def remove_card(product_id):
    with get_engine().begin() as c:
        c.execute(delete(cards).where(cards.c.product_id == product_id))
        c.execute(update(products).where(products.c.id == product_id).values(card_file=None))


def get_card_bytes(product_id) -> bytes | None:
    with get_engine().connect() as c:
        row = c.execute(select(cards.c.data).where(cards.c.product_id == product_id)).first()
    return row[0] if row else None


def materialize_card(product_id) -> str | None:
    """Zapisuje bajty karty do pliku tymczasowego i zwraca ścieżkę (do scalenia PDF)."""
    data = get_card_bytes(product_id)
    if not data:
        return None
    CARDS_DIR.mkdir(parents=True, exist_ok=True)
    p = CARDS_DIR / ("%s.pdf" % product_id)
    p.write_bytes(data)
    return str(p)


# --------------------------------------------------------------------------- #
#  Ustawienia
# --------------------------------------------------------------------------- #
def get_settings() -> dict:
    init_db()
    with get_engine().connect() as c:
        rows = dict(c.execute(select(settings_t.c.key, settings_t.c.value)).all())
    out = dict(DEFAULT_SETTINGS)
    out.update(rows)
    return out


def save_settings(d: dict):
    with get_engine().begin() as c:
        for k, v in d.items():
            exists = c.execute(select(settings_t.c.key).where(settings_t.c.key == k)).first()
            if exists:
                c.execute(update(settings_t).where(settings_t.c.key == k).values(value=str(v)))
            else:
                c.execute(insert(settings_t).values(key=k, value=str(v)))


# --------------------------------------------------------------------------- #
#  Oferty
# --------------------------------------------------------------------------- #
def next_offer_number() -> str:
    s = get_settings()
    now = datetime.now()
    key = "offer_seq_%d" % now.year
    with get_engine().begin() as c:
        row = c.execute(select(settings_t.c.value).where(settings_t.c.key == key)).first()
        seq = int(row[0]) + 1 if row else 1
        if row:
            c.execute(update(settings_t).where(settings_t.c.key == key).values(value=str(seq)))
        else:
            c.execute(insert(settings_t).values(key=key, value=str(seq)))
    return "%s/%03d/%02d/%d" % (s["offer_prefix"], seq, now.month, now.year)


def save_offer(number, client, tier, items, total_net, pdf_name, pdf_bytes,
               user_id=None, user_name="", deal_id=None) -> int:
    init_db()
    with get_engine().begin() as c:
        res = c.execute(insert(offers).values(
            number=number, ts=_now(), user_id=user_id, user_name=user_name,
            client_name=client.get("nazwa", ""), client_json=json.dumps(client, ensure_ascii=False),
            tier=tier, items_json=json.dumps(items, ensure_ascii=False, default=str),
            total_net=total_net, pdf_name=pdf_name, pdf_blob=pdf_bytes,
            pipedrive_deal_id=deal_id))
        pk = res.inserted_primary_key
        return pk[0] if pk else 0


def offers_df(limit=200, user_id=None, is_admin=False) -> pd.DataFrame:
    init_db()
    q = "SELECT id,number,ts,user_name,client_name,tier,total_net,pdf_name FROM offers"
    params = {"l": limit}
    if not is_admin and user_id is not None:
        q += " WHERE user_id = :uid"
        params["uid"] = user_id
    q += " ORDER BY id DESC LIMIT :l"
    with get_engine().connect() as c:
        return pd.read_sql_query(text(q), c, params=params)


def get_offer_pdf(offer_id) -> tuple[str, bytes] | None:
    with get_engine().connect() as c:
        row = c.execute(select(offers.c.pdf_name, offers.c.pdf_blob)
                        .where(offers.c.id == offer_id)).first()
    return (row[0], row[1]) if row and row[1] else None


# --------------------------------------------------------------------------- #
#  Użytkownicy
# --------------------------------------------------------------------------- #
def count_users() -> int:
    init_db()
    with get_engine().connect() as c:
        return c.execute(select(func.count()).select_from(users)).scalar() or 0


def get_user_by_email(email: str):
    with get_engine().connect() as c:
        row = c.execute(select(users).where(func.lower(users.c.email) == email.lower().strip())).first()
    return dict(row._mapping) if row else None


def get_user(uid: int):
    with get_engine().connect() as c:
        row = c.execute(select(users).where(users.c.id == uid)).first()
    return dict(row._mapping) if row else None


def create_user(email, name, phone, password_hash, role="handlowiec",
                pipedrive_token=None) -> int:
    init_db()
    with get_engine().begin() as c:
        res = c.execute(insert(users).values(
            email=email.strip(), name=name.strip(), phone=(phone or "").strip(),
            password_hash=password_hash, role=role, active=1, created_ts=_now(),
            pipedrive_token=(pipedrive_token or None)))
        pk = res.inserted_primary_key
        return pk[0] if pk else 0


def update_user(uid, **fields):
    if not fields:
        return
    with get_engine().begin() as c:
        c.execute(update(users).where(users.c.id == uid).values(**fields))


def users_df() -> pd.DataFrame:
    init_db()
    with get_engine().connect() as c:
        return pd.read_sql_query(
            text("SELECT id,email,name,phone,role,active,created_ts FROM users ORDER BY id"), c)


def _now():
    return datetime.now().isoformat(timespec="seconds")
